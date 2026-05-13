import argparse
import csv
import hashlib
import json
import re
import time
from copy import copy
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import urljoin, urlparse

import requests


STAYSURE_COLUMNS = [
    "id",
    "img1",
    "img2",
    "img3",
    "img4",
    "img5",
    "location",
    "size_sqft",
    "furnishing_type",
    "cleanliness_score",
    "amenities",
    "rent_price",
    "bhk",
]
EXTRA_COLUMNS = ["bathroom_attached", "title", "description", "source_url"]
IMAGE_EXTENSIONS = (".jpg", ".jpeg", ".png", ".webp")
AMENITY_KEYWORDS = {
    "AC": [" ac ", "air conditioner", "air conditioning"],
    "Attached Bathroom": ["attached bathroom", "attached washroom", "attach bathroom"],
    "Bed": [" bed ", "mattress", "cot"],
    "CCTV": ["cctv"],
    "Fan": [" fan ", "ceiling fan"],
    "Fridge": ["fridge", "refrigerator"],
    "Kitchen": ["kitchen"],
    "Lift": ["lift", "elevator"],
    "Parking": ["parking"],
    "Power Backup": ["power backup", "inverter"],
    "Security": ["security", "guard"],
    "Sofa": ["sofa"],
    "Table": ["table", "study table"],
    "Washing Machine": ["washing machine", "washer"],
    "Water Supply": ["water supply", "24 hrs water", "24 hours water", "water"],
    "Wifi": ["wifi", "wi-fi", "internet"],
}


@dataclass
class Listing:
    url: str
    title: str = ""
    description: str = ""
    location: str = "Unknown"
    size_sqft: int = 0
    furnishing_type: str = "Unknown"
    cleanliness_score: float = 5.0
    amenities: list[str] = field(default_factory=list)
    rent_price: int = 0
    bhk: str = "Unknown"
    bathroom_attached: int = 0
    image_urls: list[str] = field(default_factory=list)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Collect public OLX rental listings into the StaySure dataset format."
    )
    parser.add_argument(
        "--start-url",
        action="append",
        default=None,
        help="OLX category/search URL to collect from. Pass multiple times to merge regions/searches.",
    )
    parser.add_argument("--max-listings", type=int, default=100)
    parser.add_argument("--start-id", type=int, default=1)
    parser.add_argument("--links-input", type=Path, default=None)
    parser.add_argument("--links-output", type=Path, default=None)
    parser.add_argument("--links-only", action="store_true")
    parser.add_argument("--output-csv", type=Path, default=Path("Dataset/rooms_dataset.csv"))
    parser.add_argument("--output-xlsx", type=Path, default=None)
    parser.add_argument("--image-dir", type=Path, default=Path("Dataset/rooms"))
    parser.add_argument("--download-images", action="store_true")
    parser.add_argument("--headful", action="store_true", help="Show the browser while collecting.")
    parser.add_argument("--delay", type=float, default=1.5, help="Delay between listing pages.")
    parser.add_argument("--page-wait-ms", type=int, default=2500, help="Max wait after opening each listing page.")
    parser.add_argument(
        "--room-only",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="Keep listings that look like room/PG/bachelor rental posts.",
    )
    parser.add_argument(
        "--include-extra-columns",
        action="store_true",
        help="Append bathroom_attached, title, description, and source_url columns.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    try:
        from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
        from playwright.sync_api import sync_playwright
    except ImportError as exc:
        raise SystemExit(
            "Playwright is required for OLX collection. Install with:\n"
            "  pip install -r Dataset/requirements-scraper.txt\n"
            "  python -m playwright install chromium"
        ) from exc

    args.output_csv.parent.mkdir(parents=True, exist_ok=True)
    args.image_dir.mkdir(parents=True, exist_ok=True)

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(
            headless=not args.headful,
            args=[
                "--disable-http2",
                "--disable-quic",
                "--disable-blink-features=AutomationControlled",
            ],
        )
        context = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
                "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1366, "height": 900},
        )
        page = context.new_page()

        if args.links_input:
            links = read_links(args.links_input)[: args.max_listings]
            print(f"Loaded {len(links)} listing links from {args.links_input}")
        else:
            start_urls = normalize_start_urls(args.start_url)
            links = collect_listing_links(page, start_urls, args.max_listings, PlaywrightTimeoutError)
        if args.links_output:
            write_links(args.links_output, links)
            print(f"Wrote {len(links)} listing links to {args.links_output}")
        if args.links_only:
            browser.close()
            return
        rows = []
        for offset, url in enumerate(links):
            print(f"[{offset + 1}/{len(links)}] {url}")
            try:
                listing = collect_listing(page, url, PlaywrightTimeoutError, args.page_wait_ms)
            except Exception as exc:  # noqa: BLE001 - keep long collection jobs moving.
                print(f"  skipped: {exc}")
                continue
            if args.room_only and not is_room_listing(listing):
                print(f"  skipped: listing does not look room-focused ({listing.title})")
                continue

            listing_id = args.start_id + len(rows)
            image_names = ["", "", "", "", ""]
            if args.download_images:
                image_names = download_images(listing.image_urls[:5], args.image_dir, listing_id)
            else:
                for index, image_url in enumerate(listing.image_urls[:5]):
                    image_names[index] = image_url

            row = format_row(listing_id, listing, image_names, args.include_extra_columns)
            rows.append(row)
            time.sleep(args.delay)

        browser.close()

    write_csv(args.output_csv, rows, args.include_extra_columns)
    if args.output_xlsx:
        write_xlsx(args.output_xlsx, rows, args.include_extra_columns)
    print(f"Wrote {len(rows)} rows to {args.output_csv}")
    if args.output_xlsx:
        print(f"Wrote spreadsheet to {args.output_xlsx}")


def normalize_start_urls(value: str | list[str]) -> list[str]:
    if value is None:
        return ["https://www.olx.in/bhilai_g4059463/for-rent-houses-apartments_c1723"]
    if isinstance(value, str):
        return [value]
    return value


def read_links(path: Path) -> list[str]:
    return [line.strip() for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]


def write_links(path: Path, links: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(links) + "\n", encoding="utf-8")


def collect_listing_links(page: Any, start_urls: list[str], limit: int, timeout_error: type[Exception]) -> list[str]:
    links: list[str] = []
    seen = set()
    for start_url in start_urls:
        if len(links) >= limit:
            break
        print(f"Opening {start_url}")
        safe_goto(page, start_url, timeout=60_000)
        idle_rounds = 0

        while len(links) < limit and idle_rounds < 8:
            before = len(links)
            for href in page.eval_on_selector_all("a[href]", "(nodes) => nodes.map((node) => node.href)"):
                if not is_listing_url(href):
                    continue
                normalized = normalize_url(href)
                if normalized in seen:
                    continue
                seen.add(normalized)
                links.append(normalized)
                if len(links) >= limit:
                    break

            try:
                page.get_by_text(re.compile(r"load more|show more", re.I)).click(timeout=1200)
            except timeout_error:
                page.mouse.wheel(0, 2600)
                page.wait_for_timeout(1200)

            idle_rounds = idle_rounds + 1 if len(links) == before else 0
        print(f"  collected {len(links)} unique listing links so far")

    return links[:limit]


def collect_listing(page: Any, url: str, timeout_error: type[Exception], page_wait_ms: int = 2500) -> Listing:
    safe_goto(page, url, timeout=60_000)
    try:
        page.wait_for_load_state("networkidle", timeout=page_wait_ms)
    except timeout_error:
        pass

    raw_body_text = page.locator("body").inner_text(timeout=10_000)
    body_text = clean_text(raw_body_text)
    scripts = page.eval_on_selector_all("script", "(nodes) => nodes.map((node) => node.textContent || '')")
    meta = extract_meta(page)
    json_values = list(iter_json_values(scripts))

    text_blob = " ".join([body_text, meta.get("title", ""), meta.get("description", "")])
    listing = Listing(url=url)
    listing.title = first_non_empty(
        selector_text(page, "h1"),
        meta.get("title", ""),
        find_json_string(json_values, ["title", "name"]),
    )
    listing.description = first_non_empty(
        selector_text(page, '[data-aut-id="itemDescriptionContent"]'),
        meta.get("description", ""),
        find_json_string(json_values, ["description"]),
    )
    listing.location = first_non_empty(
        selector_text(page, '[data-aut-id="item-location"]'),
        find_location_from_text(raw_body_text),
        find_json_string(json_values, ["location", "city", "locality", "address"]),
        "Unknown",
    )
    listing.rent_price = parse_price(first_non_empty(selector_text(page, '[data-aut-id="itemPrice"]'), text_blob))
    listing.size_sqft = parse_size(text_blob)
    listing.bhk = parse_bhk(text_blob)
    listing.furnishing_type = parse_furnishing(text_blob)
    listing.bathroom_attached = 1 if re.search(r"\battach(?:ed)?\s+(bathroom|washroom)\b", text_blob, re.I) else 0
    listing.cleanliness_score = estimate_cleanliness(text_blob)
    listing.amenities = infer_amenities(text_blob)
    listing.description = redact_private_text(listing.description)
    page.mouse.wheel(0, 1200)
    page.wait_for_timeout(500)
    listing.image_urls = collect_image_urls(page, scripts, json_values, url)
    return listing


def safe_goto(page: Any, url: str, timeout: int) -> None:
    last_error: Exception | None = None
    for wait_until in ("domcontentloaded", "commit", "load"):
        try:
            page.goto(url, wait_until=wait_until, timeout=timeout)
            return
        except Exception as exc:  # noqa: BLE001
            last_error = exc
            page.wait_for_timeout(1500)
    raise RuntimeError(f"Could not open {url}: {last_error}") from last_error


def extract_meta(page: Any) -> dict[str, str]:
    return page.evaluate(
        """
        () => {
          const output = {};
          for (const element of document.querySelectorAll('meta')) {
            const key = element.getAttribute('property') || element.getAttribute('name');
            const value = element.getAttribute('content');
            if (key && value) output[key] = value;
          }
          return {
            title: document.title || output['og:title'] || '',
            description: output['description'] || output['og:description'] || '',
            image: output['og:image'] || ''
          };
        }
        """
    )


def collect_image_urls(page: Any, scripts: list[str], json_values: list[Any], base_url: str) -> list[str]:
    candidates: list[str] = []
    meta = extract_meta(page)
    if meta.get("image"):
        candidates.append(meta["image"])

    candidates.extend(collect_visible_listing_images(page))
    for value in json_values:
        if isinstance(value, str) and looks_like_image_url(value):
            candidates.append(value)
    for script in scripts:
        candidates.extend(re.findall(r"https?://[^\"'\\s]+?(?:jpg|jpeg|png|webp)", script, flags=re.I))

    normalized = []
    seen = set()
    for candidate in candidates:
        for image_candidate in split_srcset(candidate):
            absolute = normalize_url(urljoin(base_url, image_candidate))
            if should_skip_image_url(absolute):
                continue
            key = canonical_image_key(absolute)
            if key in seen or not looks_like_image_url(absolute):
                continue
            seen.add(key)
            normalized.append(absolute)
    return normalized


def collect_visible_listing_images(page: Any) -> list[str]:
    images = page.eval_on_selector_all(
        "img",
        """
        (nodes) => nodes.map((node) => {
          const rect = node.getBoundingClientRect();
          return {
            src: node.currentSrc || node.src || "",
            srcset: node.srcset || node.getAttribute("data-srcset") || "",
            dataSrc: node.getAttribute("data-src") || "",
            alt: node.alt || "",
            width: node.naturalWidth || rect.width || 0,
            height: node.naturalHeight || rect.height || 0,
            top: rect.top,
            left: rect.left
          };
        })
        """,
    )
    listing_images = []
    for image in images:
        if not looks_like_listing_image_node(image):
            continue
        listing_images.extend([image.get("src", ""), image.get("srcset", ""), image.get("dataSrc", "")])
    return listing_images


def looks_like_listing_image_node(image: dict[str, Any]) -> bool:
    src = str(image.get("src", ""))
    alt = str(image.get("alt", "")).lower()
    width = float(image.get("width") or 0)
    height = float(image.get("height") or 0)
    if should_skip_image_url(src):
        return False
    if "apollo.olx" not in urlparse(src).netloc.lower():
        return False
    if width < 250 or height < 250:
        return False
    if any(term in alt for term in ["cars", "bikes", "properties", "facebook", "instagram", "youtube"]):
        return False
    return True


def split_srcset(value: str) -> Iterable[str]:
    for part in str(value).split(","):
        url = part.strip().split(" ")[0]
        if url:
            yield url


def should_skip_image_url(url: str) -> bool:
    lowered = url.lower()
    skip_terms = [
        "logo",
        "icon",
        "playstore",
        "appstore",
        "favicon",
        "/external/base/img/",
        "panamera",
        "alias-",
    ]
    return any(term in lowered for term in skip_terms)


def canonical_image_key(url: str) -> str:
    match = re.search(r"/v1/files/([^/]+)/image", urlparse(url).path)
    if match:
        return match.group(1)
    return normalize_url(url)


def looks_like_image_url(url: str) -> bool:
    parsed = urlparse(url)
    path = parsed.path.lower()
    host = parsed.netloc.lower()
    if any(ext in path for ext in IMAGE_EXTENSIONS):
        return True
    return "apollo.olx" in host and "/v1/files/" in path


def normalize_url(url: str) -> str:
    parsed = urlparse(url)
    return parsed._replace(fragment="").geturl()


def clean_text(value: str) -> str:
    return re.sub(r"\s+", " ", value or "").strip()


def redact_private_text(value: str) -> str:
    value = re.sub(r"\b\d[\d _.-]{7,}\d\b", "[redacted]", value)
    return clean_text(value)


def first_non_empty(*values: str) -> str:
    for value in values:
        value = clean_text(value)
        if value:
            return value
    return ""


def iter_json_values(scripts: Iterable[str]) -> Iterable[Any]:
    for script in scripts:
        for raw in extract_json_candidates(script):
            try:
                parsed = json.loads(raw)
            except json.JSONDecodeError:
                continue
            yield from walk_json(parsed)


def extract_json_candidates(script: str) -> Iterable[str]:
    stripped = script.strip()
    if not stripped:
        return
    if stripped.startswith("{") or stripped.startswith("["):
        yield stripped
    for match in re.finditer(r"<script[^>]+type=[\"']application/ld\+json[\"'][^>]*>(.*?)</script>", script, re.I | re.S):
        yield match.group(1)
    for match in re.finditer(r"window\.__[A-Z0-9_]+__\s*=\s*({.*?});", script, re.S):
        yield match.group(1)


def walk_json(value: Any) -> Iterable[Any]:
    yield value
    if isinstance(value, dict):
        for item in value.values():
            yield from walk_json(item)
    elif isinstance(value, list):
        for item in value:
            yield from walk_json(item)


def find_json_string(values: list[Any], keys: list[str]) -> str:
    key_set = {key.lower() for key in keys}
    for value in values:
        if not isinstance(value, dict):
            continue
        for key, item in value.items():
            if str(key).lower() not in key_set:
                continue
            if isinstance(item, str) and item.strip():
                return clean_text(item)
            if isinstance(item, dict):
                nested = " ".join(str(part) for part in item.values() if isinstance(part, str))
                if nested.strip():
                    return clean_text(nested)
    return ""


def selector_text(page: Any, selector: str) -> str:
    try:
        locator = page.locator(selector).first
        if locator.count() == 0:
            return ""
        return clean_text(locator.inner_text(timeout=1200))
    except Exception:  # noqa: BLE001
        return ""


def format_row(listing_id: int, listing: Listing, images: list[str], include_extra: bool) -> dict[str, Any]:
    row: dict[str, Any] = {
        "id": listing_id,
        "img1": images[0],
        "img2": images[1],
        "img3": images[2],
        "img4": images[3],
        "img5": images[4],
        "location": listing.location,
        "size_sqft": listing.size_sqft,
        "furnishing_type": listing.furnishing_type,
        "cleanliness_score": listing.cleanliness_score,
        "amenities": ", ".join(sorted(set(listing.amenities))),
        "rent_price": listing.rent_price,
        "bhk": listing.bhk,
    }
    if include_extra:
        row.update(
            {
                "bathroom_attached": listing.bathroom_attached,
                "title": listing.title,
                "description": listing.description,
                "source_url": listing.url,
            }
        )
    return row


def write_csv(path: Path, rows: list[dict[str, Any]], include_extra: bool) -> None:
    columns = STAYSURE_COLUMNS + (EXTRA_COLUMNS if include_extra else [])
    with path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(path: Path, rows: list[dict[str, Any]], include_extra: bool) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise SystemExit(
            "openpyxl is required for --output-xlsx. Install with:\n"
            "  pip install -r Dataset/requirements-scraper.txt"
        ) from exc

    columns = STAYSURE_COLUMNS + (EXTRA_COLUMNS if include_extra else [])
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(columns)
    for row in rows:
        sheet.append([row.get(column, "") for column in columns])

    for cell in sheet[1]:
        font = copy(cell.font)
        font.bold = True
        cell.font = font
    for column_cells in sheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(
            max(max_len + 2, 10),
            36,
        )
    sheet.freeze_panes = "A2"
    workbook.save(path)


def download_images(image_urls: list[str], image_dir: Path, listing_id: int) -> list[str]:
    names = ["", "", "", "", ""]
    for index, image_url in enumerate(image_urls[:5], start=1):
        suffix = image_suffix(image_url)
        name = f"room_{listing_id}_img{index}{suffix}"
        path = image_dir / name
        try:
            response = requests.get(
                image_url,
                timeout=20,
                headers={"User-Agent": "Mozilla/5.0"},
            )
            response.raise_for_status()
            path.write_bytes(response.content)
            names[index - 1] = name
        except requests.RequestException as exc:
            print(f"  image skipped: {exc}")
    return names


def image_suffix(url: str) -> str:
    suffix = Path(urlparse(url).path).suffix.lower()
    if suffix in IMAGE_EXTENSIONS:
        return suffix
    digest = hashlib.sha1(url.encode("utf-8")).hexdigest()[:8]
    return f"_{digest}.jpg"


def parse_price(text: str) -> int:
    matches = re.findall(r"(?:Rs\.?|INR|₹)\s*([0-9][0-9,]*(?:\.\d+)?)", text, flags=re.I)
    if not matches:
        matches = re.findall(r"\b([0-9][0-9,]{3,})\b", text)
    for match in matches:
        value = int(float(match.replace(",", "")))
        if 500 <= value <= 1_000_000:
            return value
    return 0


def parse_size(text: str) -> int:
    match = re.search(r"\b([0-9]{2,5})\s*(?:sq\.?\s*ft|sqft|square feet)\b", text, flags=re.I)
    if not match:
        return 0
    return int(match.group(1))


def parse_bhk(text: str) -> str:
    match = re.search(r"\b([1-9])\s*(?:BHK|RK)\b", text, flags=re.I)
    if not match:
        return "Unknown"
    return match.group(1)


def parse_furnishing(text: str) -> str:
    lowered = text.lower()
    if "semi furnished" in lowered or "semi-furnished" in lowered:
        return "Semi-furnished"
    if "unfurnished" in lowered:
        return "Unfurnished"
    if "furnished" in lowered or "fully furnished" in lowered:
        return "furnished"
    return "Unknown"


def estimate_cleanliness(text: str) -> float:
    lowered = text.lower()
    score = 5.0
    positive = [
        "newly constructed",
        "new construction",
        "well maintained",
        "clean",
        "neat",
        "premium",
        "renovated",
        "good condition",
    ]
    negative = ["old", "needs repair", "renovation required", "dirty", "damaged"]
    score += sum(0.8 for word in positive if word in lowered)
    score -= sum(1.0 for word in negative if word in lowered)
    return round(max(1.0, min(score, 10.0)), 1)


def infer_amenities(text: str) -> list[str]:
    padded = f" {text.lower()} "
    amenities = []
    for label, keywords in AMENITY_KEYWORDS.items():
        if any(keyword in padded for keyword in keywords):
            amenities.append(label)
    return amenities


def find_location_from_text(text: str) -> str:
    compact = clean_text(text)
    patterns = [
        r"Posted in\s+([^,]+,\s*Bhilai,\s*Chhattisgarh)",
        r"([A-Za-z0-9 .'-]+,\s*Bhilai,\s*Chhattisgarh)",
    ]
    for pattern in patterns:
        match = re.search(pattern, compact, re.I)
        if match:
            return clean_text(match.group(1))

    lines = [line.strip() for line in text.splitlines() if line.strip()]
    for line in lines:
        if (
            re.search(r"\b(Bhilai|Durg|Raipur|Chhattisgarh)\b", line, re.I)
            and len(line) <= 120
            and not re.search(r"\b(OLX|Wishlist|Login|SELL|REPORT|POPULAR)\b", line)
        ):
            return clean_text(line)
    return ""


def is_room_listing(listing: Listing) -> bool:
    text = f"{listing.title} {listing.description} {' '.join(listing.amenities)}".lower()
    room_terms = [
        "room",
        "single room",
        "bachelor",
        "bachelors",
        "pg",
        "paying guest",
        "hostel",
        "student",
        "students",
    ]
    if any(term in text for term in room_terms):
        return True
    if listing.bhk in {"1", "1.0"} and listing.size_sqft and listing.size_sqft <= 800:
        return True
    return False


def is_listing_url(url: str) -> bool:
    parsed = urlparse(url)
    return "olx.in" in parsed.netloc and "/item/" in parsed.path


if __name__ == "__main__":
    main()
