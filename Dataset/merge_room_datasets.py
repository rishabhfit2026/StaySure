import argparse
import csv
from pathlib import Path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Append compatible StaySure room dataset CSV files.")
    parser.add_argument("--base-csv", type=Path, required=True)
    parser.add_argument("--append-csv", type=Path, required=True)
    parser.add_argument("--output-csv", type=Path, required=True)
    parser.add_argument("--output-xlsx", type=Path, default=None)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    columns, rows = read_rows(args.base_csv)
    append_columns, append_rows = read_rows(args.append_csv)
    if append_columns != columns:
        raise ValueError(f"CSV columns do not match: {append_columns} != {columns}")

    rows.extend(append_rows)
    write_csv(args.output_csv, columns, rows)
    if args.output_xlsx:
        write_xlsx(args.output_xlsx, columns, rows)
    print(f"Merged {len(rows)} rows into {args.output_csv}")


def read_rows(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open(newline="", encoding="utf-8") as file:
        reader = csv.DictReader(file)
        columns = list(reader.fieldnames or [])
        return columns, [{column: row.get(column, "") for column in columns} for row in reader]


def write_csv(path: Path, columns: list[str], rows: list[dict[str, str]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(path: Path, columns: list[str], rows: list[dict[str, str]]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise SystemExit("openpyxl is required to write XLSX output.") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(columns)
    for row in rows:
        sheet.append([row.get(column, "") for column in columns])

    for cell in sheet[1]:
        font = cell.font.copy()
        font.bold = True
        cell.font = font
    for column_cells in sheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_len + 2, 10), 36)
    sheet.freeze_panes = "A2"
    workbook.save(path)


if __name__ == "__main__":
    main()
