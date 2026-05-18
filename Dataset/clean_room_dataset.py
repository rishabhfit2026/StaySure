import argparse
import csv
from pathlib import Path


IMAGE_COLUMNS = ["img1", "img2", "img3", "img4", "img5"]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Fill missing StaySure room dataset metadata fields.")
    parser.add_argument("--input-csv", type=Path, required=True)
    parser.add_argument("--output-csv", type=Path, required=True)
    parser.add_argument("--output-xlsx", type=Path, default=None)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    columns, rows = read_rows(args.input_csv)
    for row in rows:
        clean_row(row)
    write_csv(args.output_csv, columns, rows)
    if args.output_xlsx:
        write_xlsx(args.output_xlsx, columns, rows)
    print(f"Cleaned {len(rows)} rows into {args.output_csv}")


def read_rows(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open(newline="", encoding="utf-8") as file:
        reader = csv.DictReader(file)
        columns = list(reader.fieldnames or [])
        return columns, [{column: row.get(column, "") for column in columns} for row in reader]


def clean_row(row: dict[str, str]) -> None:
    row["location"] = normalized_text(row.get("location"), "Unknown, Chhattisgarh")
    if "chhattisgarh" not in row["location"].lower():
        row["location"] = f"{row['location']}, Chhattisgarh"

    row["bhk"] = normalize_bhk(row.get("bhk"))
    row["size_sqft"] = str(normalize_size(row.get("size_sqft"), row["bhk"]))
    row["furnishing_type"] = normalized_text(row.get("furnishing_type"), "Unknown")
    row["cleanliness_score"] = normalized_text(row.get("cleanliness_score"), "5.0")
    row["amenities"] = normalized_text(row.get("amenities"), "None")
    row["rent_price"] = normalized_text(row.get("rent_price"), "0")


def normalized_text(value: object, default: str) -> str:
    text = str(value or "").strip()
    if not text or text.lower() in {"nan", "none", "null"}:
        return default
    return text


def normalize_bhk(value: object) -> str:
    text = normalized_text(value, "1")
    if text.lower() == "unknown":
        return "1"
    return text


def normalize_size(value: object, bhk: str) -> int:
    try:
        size = int(float(str(value or "0").replace(",", "")))
    except ValueError:
        size = 0
    if 80 <= size <= 5000:
        return size
    defaults = {"1": 300, "2": 650, "3": 1000, "4": 1400}
    return defaults.get(str(bhk), 300)


def write_csv(path: Path, columns: list[str], rows: list[dict[str, str]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(path: Path, columns: list[str], rows: list[dict[str, str]]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise SystemExit("openpyxl is required to write XLSX output.") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(columns)
    for row in rows:
        sheet.append([row.get(column, "") for column in columns])

    for cell in sheet[1]:
        font = cell.font.copy()
        font.bold = True
        cell.font = font
    for column_cells in sheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_len + 2, 10), 36)
    sheet.freeze_panes = "A2"
    workbook.save(path)


if __name__ == "__main__":
    main()
